# 当前代码可运行，检验基本可行，后续基于此做调整，目前使用此段代码
import pandas as pd
import warnings
import logging
from openpyxl.styles import PatternFill
from pathlib import Path

CONFIG = {
    # 输入文件路径
    "input_files": {
        "rent": Path(r"D:\Filelist\Python_list\File_list\汇总的租金收入表v1.csv"),
        "asset": Path(r"D:\Filelist\Python_list\历史系统数据\资产余额表06.30_处理后（保留格式）.xlsx")
    },
    # 输出文件路径
    "output_file": Path(r"D:\Filelist\Python_list\Vintage15+.商用车.损失不担.去重组.0630_拨备率.xlsx"),
    # 经销商筛选条件
    "dealer_filters": [],
    # 业务参数
    "business_params": {
        "overdue_dpd_threshold": 16,
        "first_period_num": 1,
        "roll_rate_bins": [0, 1, 31, 61, 91, 121, 151, 181, float('inf')],           # 逾期分箱边界（左闭右开，与教程完全一致）
        # 分箱标签
        "roll_rate_labels": ["M0(正常)", "M1(1-30天)", "M2(31-60天)", "M3(61-90天)",
                             "M4(91-120天)", "M5(121-150天)", "M6(151-180天)", "M7(181天+)"],
        # 迁移档位对应关系
        "roll_rate_pairs": [
            ("M0(正常)", "M1(1-30天)"),
            ("M1(1-30天)", "M2(31-60天)"),
            ("M2(31-60天)", "M3(61-90天)"),
            ("M3(61-90天)", "M4(91-120天)"),
            ("M4(91-120天)", "M5(121-150天)"),
            ("M5(121-150天)", "M6(151-180天)"),
            ("M6(151-180天)", "M7(181天+)")
        ]
    },
    # 必要列配置
    "required_columns": {
        "merge": ["合同号", "经销商名称", "客户名称", "起租日", "期号", "租金结算日期", "结清日期", "未偿还本金",
                  "放款金额", "业务类别", "业务模式"],
        "preprocess": ["合同号", "期号", "放款金额", "起租日", "逾期天数（DPD）", "未偿还本金"]
    }
}

warnings.filterwarnings("ignore", category=UserWarning, module='openpyxl.styles.stylesheet')


def init_logger():
    """初始化日志配置，同时输出到控制台和文件"""
    log_format = "%(asctime)s - %(levelname)s - %(module)s - %(message)s"
    logging.basicConfig(
        level=logging.INFO,
        format=log_format,
        handlers=[
            logging.FileHandler("vintage_analysis.log", encoding="utf-8"),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger(__name__)


# 初始化日志对象
logger = init_logger()


# 1. 数据提取
# ============================================
def merge_excel_files(analysis_date):
    """合并租金收入表和资产余额表，处理日期字段并计算逾期天数（DPD）
    :param analysis_date: 分析基准日（上月末pd.Timestamp），用于统一计算DPD
    """
    try:
        # 读取租金收入表
        rent_df = pd.read_csv(
            CONFIG["input_files"]["rent"],
            usecols=['合同号', '经销商名称', '起租日', '期号', '租金结算日期', '结清日期', '未偿还本金', '业务类别',
                     '是否结清标志']
        )
        logger.info(f"成功读取租金收入表，数据行数：{len(rent_df)}")

        # 读取资产余额表
        asset_df = pd.read_excel(
            CONFIG["input_files"]["asset"],
            usecols=['合同号', '客户名称', '放款金额', '业务模式', '大区', '业务来源']
        )
        logger.info(f"成功读取资产余额表，数据行数：{len(asset_df)}")

        # 合并数据
        merged_df = pd.merge(rent_df, asset_df, on='合同号', how='inner')
        logger.info(f"两表inner合并后，数据行数：{len(merged_df)}")

        # 1. 标准化日期字段（同原逻辑，不变）
        merged_df['租金结算日期'] = pd.to_datetime(merged_df['租金结算日期'], errors='coerce')
        merged_df['结清日期'] = pd.to_datetime(merged_df['结清日期'], errors='coerce')

        # 2. 获取程序运行的当前时间（仅保留日期，去掉时分秒，避免跨天误差）
        current_date = analysis_date.normalize()  # 例如：2025-10-02 00:00:00

        # 向量化计算逾期天数（DPD）
        merged_df['逾期天数（DPD）'] = 0  # 默认未逾期

        # 场景1：已结清合同，用结清日期计算
        mask_settled = merged_df['结清日期'].notna() & merged_df['租金结算日期'].notna()
        merged_df.loc[mask_settled, '逾期天数（DPD）'] = (
                merged_df.loc[mask_settled, '结清日期'] - merged_df.loc[mask_settled, '租金结算日期']
        ).dt.days

        # 场景2：未结清且已过结算日的合同，用当前日期计算
        mask_unsettled = merged_df['结清日期'].isna() & merged_df['租金结算日期'].notna()
        mask_overdue = mask_unsettled & (current_date > merged_df['租金结算日期'])
        merged_df.loc[mask_overdue, '逾期天数（DPD）'] = (
                current_date - merged_df.loc[mask_overdue, '租金结算日期']
        ).dt.days

        # 兜底：DPD不能为负（提前结清、未到还款日）
        merged_df['逾期天数（DPD）'] = merged_df['逾期天数（DPD）'].clip(lower=0)
        # 结算日期为空的记录标记为缺失
        merged_df.loc[merged_df['租金结算日期'].isna(), '逾期天数（DPD）'] = pd.NA

        return merged_df
    except FileNotFoundError as e:
        logger.error(f"文件未找到: {e}", exc_info=True)
    except KeyError as e:
        logger.error(f"列名错误: {e}", exc_info=True)
    except Exception as e:
        logger.error(f"数据提取未知错误: {e}", exc_info=True)
    return None


# ============================================
# 2. 数据校验
# ============================================
def validate_data(df):
    """数据合理性校验，返回校验通过的数据并记录异常"""
    df_valid = df.copy()
    invalid_records = []

    # 数值型字段校验
    if (df_valid['放款金额'] <= 0).any():
        invalid_cnt = (df_valid['放款金额'] <= 0).sum()
        invalid_records.append(f"放款金额≤0的记录数：{invalid_cnt}")
        df_valid = df_valid[df_valid['放款金额'] > 0]

    if (df_valid['期号'] <= 0).any():
        invalid_cnt = (df_valid['期号'] <= 0).sum()
        invalid_records.append(f"期号≤0的记录数：{invalid_cnt}")
        df_valid = df_valid[df_valid['期号'] > 0]

    if (df_valid['未偿还本金'] < 0).any():
        invalid_cnt = (df_valid['未偿还本金'] < 0).sum()
        invalid_records.append(f"未偿还本金<0的记录数：{invalid_cnt}")
        df_valid = df_valid[df_valid['未偿还本金'] >= 0]

    # 日期逻辑校验
    df_valid['起租日'] = pd.to_datetime(df_valid['起租日'], errors='coerce')
    df_valid['租金结算日期'] = pd.to_datetime(df_valid['租金结算日期'], errors='coerce')
    date_invalid = df_valid[(df_valid['起租日'].notna()) & (df_valid['租金结算日期'].notna()) &
                            (df_valid['起租日'] > df_valid['租金结算日期'])]
    if len(date_invalid) > 0:
        invalid_records.append(f"起租日晚于结算日期的记录数：{len(date_invalid)}")
        df_valid = df_valid.drop(date_invalid.index)

    # 输出校验结果日志
    if invalid_records:
        logger.warning("数据校验发现异常：")
        for record in invalid_records:
            logger.warning(f"  - {record}")
    else:
        logger.info("数据校验通过，无异常记录")

    return df_valid


# ============================================
# 3. 数据预处理
# ============================================
def preprocess_data(df, filter_params=None):
    """数据预处理，支持多维度筛选并计算逾期相关字段"""
    # 数据合理性校验
    df = validate_data(df)

    # 多维度筛选
    if filter_params:
        for filter_col, filter_conditions in filter_params.items():
            if filter_col not in df.columns:
                logger.warning(f"筛选列{filter_col}不存在，跳过该维度筛选")
                continue

            mask = pd.Series([True] * len(df), index=df.index)
            for condition in filter_conditions:
                if 'not ' in condition:
                    keyword = condition.replace('not ', '').strip()
                    mask &= ~df[filter_col].astype(str).str.contains(keyword, na=False, regex=False)
                    logger.info(f"筛选{filter_col}：不包含{keyword}")
                else:
                    keyword = condition.strip()
                    mask &= df[filter_col].astype(str).str.contains(keyword, na=False, regex=False)
                    logger.info(f"筛选{filter_col}：包含{keyword}")

            df = df[mask]
            logger.info(f"{filter_col}筛选后，数据行数：{len(df)}")

    # 必要列检查
    required_columns = CONFIG["required_columns"]["preprocess"]
    for col in required_columns:
        if col not in df.columns:
            raise KeyError(f"数据集中缺少必要列: {col}")

    # 数据清洗
    # 同一合同同一期号存在多条记录时，按业务状态取最终账务结果
    # 优先级：ADV（提前结清/核销调整） > Y（正常结清） > 空（未结清）
    df['状态优先级'] = df['是否结清标志'].map({'ADV': 2, 'Y': 1}).fillna(0)
    # 排序：同合同同期号，优先级高的排后面；优先级相同时，结清日期晚的排后面
    df = df.sort_values(['合同号', '期号', '状态优先级', '结清日期'])
    # 分组取最后一整行，严格保留该行的所有字段，不自动跳过空值
    df = df.groupby(['合同号', '期号'], as_index=False).tail(1)
    # 删除临时辅助列
    df = df.drop(columns=['状态优先级'])
    # 重置索引（可选，保持索引连续美观）
    df = df.reset_index(drop=True)
    df['起租日'] = pd.to_datetime(df['起租日'])
    df['Vintage'] = df['起租日'].dt.to_period('M').astype(str)

    # 逾期标记
    overdue_threshold = CONFIG["business_params"]["overdue_dpd_threshold"]
    df['is_overdue'] = df['逾期天数（DPD）'].apply(
        lambda x: 1 if (pd.notna(x) and x >= overdue_threshold) else 0
    )
    df_sorted = df.sort_values(['合同号', '期号'])

    # 合同级信息计算
    contract_is_bad = df_sorted.groupby('合同号')['is_overdue'].max().reset_index()
    contract_is_bad.rename(columns={'is_overdue': 'is_bad_asset'}, inplace=True)

    first_overdue = df_sorted[df_sorted['is_overdue'] == 1].groupby('合同号').first().reset_index()
    first_overdue = first_overdue[['合同号', '期号', '未偿还本金']].rename(
        columns={'期号': '首次逾期期号', '未偿还本金': '首次逾期未偿还本金'}
    )

    # 合并合同级信息
    df = pd.merge(df, contract_is_bad, on='合同号', how='left')
    df = pd.merge(df, first_overdue, on='合同号', how='left')
    df['首次逾期未偿还本金'] = df['首次逾期未偿还本金'].fillna(0)
    df['首次逾期期号'] = df['首次逾期期号'].fillna(0).astype(int)

    # 替换原来的apply写法
    df['当期逾期本金'] = 0.0
    mask_after_overdue = df['期号'] >= df['首次逾期期号']
    df.loc[mask_after_overdue, '当期逾期本金'] = df.loc[mask_after_overdue, '首次逾期未偿还本金'].values

    return df


def fill_missing_periods(df, default_total_period=48):
    """
    仅对坏资产补全后续期数：假设坏资产默认最大期数为48期，好资产即使提前结清也不补全
    适用场景：无总期数相关字段，且仅需追踪坏资产的完整逾期表现
    参数：default_total_period - 坏资产默认最大期数，默认48期
    """
    logger.warning(f"未获取到总期数相关数据，采用简便算法：仅对坏资产补全至默认{default_total_period}期")

    # 1. 计算每个合同的“实际结清期”（现有记录中的最大期号）
    contract_last_period = df.groupby('合同号')['期号'].max().reset_index()
    contract_last_period.rename(columns={'期号': '实际结清期'}, inplace=True)

    # 2. 提取合同级关键信息（是否坏资产、首次逾期本金）
    contract_info = df.groupby('合同号').agg(
        是否坏资产=('is_bad_asset', 'max'),  # 1=坏资产，0=好资产
        首次逾期未偿还本金=('首次逾期未偿还本金', 'max')
    ).reset_index()

    # 合并“实际结清期”到合同信息中
    contract_info = pd.merge(contract_info, contract_last_period, on='合同号', how='left')

    # 3. 识别需要补全的合同（核心修改：仅坏资产且实际结清期 < 默认期数）
    need_fill_contracts = contract_info[
        (contract_info['是否坏资产'] == 1) &  # 新增：仅坏资产需要补全
        (contract_info['实际结清期'] < default_total_period) &  # 提前结清
        (contract_info['实际结清期'].notna())  # 排除异常值
        ]
    logger.info(f"需补全后续期数的坏资产合同数：{len(need_fill_contracts)}（实际结清期＜{default_total_period}期）")

    # 新增：统计“好资产且提前结清”的合同（明确说明不补全）
    good_asset_early_settle = contract_info[
        (contract_info['是否坏资产'] == 0) &  # 好资产
        (contract_info['实际结清期'] < default_total_period) &  # 提前结清
        (contract_info['实际结清期'].notna())
        ]
    if len(good_asset_early_settle) > 0:
        logger.info(f"好资产提前结清的合同数：{len(good_asset_early_settle)}（不补全后续期数）")

    # 4. 统计“坏资产但已还满默认期数”的合同（无需补全）
    bad_asset_full_period = contract_info[
        (contract_info['是否坏资产'] == 1) &  # 坏资产
        (contract_info['实际结清期'] >= default_total_period) &  # 已还满默认期数
        (contract_info['实际结清期'].notna())
        ]
    if len(bad_asset_full_period) > 0:
        logger.info(f"坏资产已还满{default_total_period}期的合同数：{len(bad_asset_full_period)}（无需补全）")

    # 5. 生成补全记录（仅针对坏资产）
    filled_records = []
    for _, contract in need_fill_contracts.iterrows():
        contract_id = contract['合同号']
        last_period = int(contract['实际结清期'])
        bad_principal = contract['首次逾期未偿还本金']  # 坏资产的逾期本金基数

        # 获取合同基础信息（复制非期数相关字段）
        base_info = df[df['合同号'] == contract_id].iloc[0].to_dict()

        # 补全从“实际结清期+1”到“默认最大期数”的所有期数
        for period in range(last_period + 1, default_total_period + 1):
            filled_record = base_info.copy()
            filled_record.update({
                '期号': period,
                '结清状态': '坏资产补全（默认48期）',  # 明确标记是坏资产的补全记录
                '租金结算日期': pd.NaT,
                '结清日期': base_info.get('结清日期'),
                '当期逾期本金': bad_principal,  # 坏资产持续计逾期本金
                'is_overdue': 1,  # 坏资产补全期仍视为逾期
                '逾期天数（DPD）': 999,  # 标记持续逾期
                '未偿还本金': bad_principal  # 坏资产补全期未偿还本金不变
            })
            filled_records.append(filled_record)

    # 6. 合并补全记录到原始数据
    if filled_records:
        filled_df = pd.DataFrame(filled_records)[df.columns]
        df = pd.concat([df, filled_df], ignore_index=True)
        logger.info(f"坏资产补全完成：新增{len(filled_records)}条记录（覆盖{len(need_fill_contracts)}个合同）")
    else:
        logger.info("无需要补全的坏资产合同")

    return df


# ============================================
# 4. 构建Vintage-MOB矩阵
# ============================================
def build_vintage_matrix(df):
    """
    构建Vintage-MOB矩阵（基于合同级放款金额，确保每个合同只统计一次）
    """
    # 1. 提取合同级唯一信息：每个合同的放款金额和所属Vintage（去重，避免重复计算）
    # 每个合同只保留一条记录（因放款金额和Vintage对同一合同是固定值）
    contract_level = df.drop_duplicates(subset=['合同号'])[
        ['合同号', 'Vintage', '放款金额']
    ]
    logger.info(f"合同级去重后的数据量：{len(contract_level)}条（每个合同一条）")

    # 2. 按Vintage汇总总放款金额（每个Vintage的总放款=该组下所有合同的放款金额之和）
    vintage_total = contract_level.groupby('Vintage')['放款金额'].sum().reset_index(
        name='总放款金额'
    )
    # 处理放款金额为0的极端情况（避免后续除零）
    vintage_total['总放款金额'] = vintage_total['总放款金额'].replace(0, pd.NA)

    # 3. 按Vintage和期号聚合逾期本金
    vintage_mob = df.groupby(['Vintage', '期号']).agg(
        逾期本金=('当期逾期本金', 'sum')
    ).reset_index()

    # 4. 合并总放款金额并计算逾期率
    vintage_mob = vintage_mob.merge(vintage_total, on='Vintage', how='left')
    # 逾期率=逾期本金/总放款金额（总放款为NA时逾期率设为0）
    vintage_mob['逾期率'] = vintage_mob.apply(
        lambda row: row['逾期本金'] / row['总放款金额'] if pd.notna(row['总放款金额']) else 0,
        axis=1
    )
    # 处理异常值（如总放款为0导致的inf）
    vintage_mob['逾期率'] = vintage_mob['逾期率'].replace([float('inf'), -float('inf')], 0)

    logger.info(
        f"Vintage矩阵构建完成：{vintage_mob['Vintage'].nunique()}个Vintage × {vintage_mob['期号'].nunique()}个MOB"
    )
    return vintage_mob


# ============================================
# 5. 过滤无效MOB（未演化到的期数）
# ============================================
def filter_invalid_mob(vintage_mob_matrix, analysis_date):
    """过滤Vintage矩阵中未演化到的无效MOB期数"""
    # 修正：先处理NaT，再转换日期（避免NaT-01）
    # 1. 标记Vintage列非NaT的有效数据
    valid_mask = vintage_mob_matrix['Vintage'].notna()

    # 2. 仅对有效数据拼接-01并转日期（无效数据设为NaT）
    vintage_mob_matrix.loc[valid_mask, 'Vintage_date'] = pd.to_datetime(
        vintage_mob_matrix.loc[valid_mask, 'Vintage'].astype(str) + '-01',
        format='%Y-%m-%d',  # 明确日期格式，提高转换效率
        errors='coerce'  # 极端异常值转为NaT，不中断程序
    )
    vintage_mob_matrix.loc[~valid_mask, 'Vintage_date'] = pd.NaT  # 无效数据设为NaT

    # 3. 删除Vintage_date为空的无效行（避免后续计算出错）
    vintage_mob_matrix = vintage_mob_matrix.dropna(subset=['Vintage_date'])

    # 新增：过滤未来Vintage（Vintage日期晚于分析基准日）
    vintage_mob_matrix = vintage_mob_matrix[vintage_mob_matrix['Vintage_date'] <= analysis_date]

    # 计算每个Vintage的最大有效MOB
    vintage_mob_matrix['month_diff'] = (
            (analysis_date.year - vintage_mob_matrix['Vintage_date'].dt.year) * 12
            + (analysis_date.month - vintage_mob_matrix['Vintage_date'].dt.month)
    )

    # 修正：month_diff<0时设为0（无有效MOB），再clip（避免0）
    vintage_mob_matrix['max_valid_mob'] = vintage_mob_matrix['month_diff'].clip(lower=0)

    # 筛选有效数据
    valid_mob_matrix = vintage_mob_matrix[
        (vintage_mob_matrix['期号'] <= vintage_mob_matrix['max_valid_mob'])
        & (vintage_mob_matrix['max_valid_mob'] > 0)
        ].copy()

    # 清理临时列
    valid_mob_matrix = valid_mob_matrix.drop(columns=['Vintage_date', 'month_diff', 'max_valid_mob'])
    logger.info(f"有效MOB筛选完成：原始记录数{len(vintage_mob_matrix)} → 有效记录数{len(valid_mob_matrix)}")

    # 打印各Vintage的有效MOB范围
    valid_mob_summary = valid_mob_matrix.groupby('Vintage').agg(
        最小有效MOB=('期号', 'min'),
        最大有效MOB=('期号', 'max')
    ).reset_index()
    logger.info("各Vintage有效MOB范围：")
    for _, row in valid_mob_summary.iterrows():
        logger.info(f"  - {row['Vintage']}: MOB{row['最小有效MOB']}~MOB{row['最大有效MOB']}")

    return valid_mob_matrix


# ============================================
# 6. 月度逾期余额表（对应教程表1）
# ============================================
def build_monthly_balance(df, analysis_date, export_month=None, export_path="月末核对明细.xlsx"):
    """
    统计每个月末各逾期等级的未偿还本金余额，生成资产金额分布表
    完全对应教程「表1-资产金额分布统计」
    """
    bins = CONFIG["business_params"]["roll_rate_bins"]
    labels = CONFIG["business_params"]["roll_rate_labels"]

    df_real = df.copy()
    contract_base = df_real.groupby('合同号').agg(
        放款金额=('放款金额', 'first'),
        起租日=('起租日', 'min')
    ).reset_index()
    contract_base = contract_base.set_index('合同号')
    contract_base = contract_base[contract_base['起租日'] >= pd.Timestamp('2023-01-01')]  # 仅保留2023年及以后起租的合同

    # 1. 生成观察月份范围：从最早起租日到分析基准日
    min_date = contract_base['起租日'].min()
    if pd.isna(min_date):
        logger.error("无有效起租日期，无法生成月度余额表")
        return pd.DataFrame()

    month_range = pd.period_range(start=min_date.to_period('M'), end=analysis_date.to_period('M'), freq='M')
    month_ends = [pd.Timestamp(m.end_time) for m in month_range]

    monthly_balance_list = []

    for month_end in month_ends:
        month_str = month_end.strftime('%Y-%m')

        mask_launched = contract_base['起租日'] <= month_end
        launched_contracts = contract_base[mask_launched].copy()

        if launched_contracts.empty:
            balance_series = pd.Series(0, index=labels, name=month_str)
            monthly_balance_list.append(balance_series)
            continue

        # 2. 计算每个合同的剩余本金（全量合同都算，最后用本金>0筛选在贷）
        # 2.1 筛选截至月末已实际结清的期数（仅结清的期数才扣减本金）
        mask_settled = (
                df_real['结清日期'].notna()
                & (df_real['结清日期'] <= month_end)
                & df_real['合同号'].isin(launched_contracts.index)
        )
        df_settled = df_real[mask_settled]

        if not df_settled.empty:
            df_settled_sorted = df_settled.sort_values('期号')
            settled_principal = df_settled_sorted.groupby('合同号')['未偿还本金'].last()
        else:
            settled_principal = pd.Series(dtype='float64')

        # 2.2 合并计算剩余本金：已结清取最后一期剩余，未结清取放款金额
        launched_contracts['剩余本金'] = settled_principal
        launched_contracts['剩余本金'] = launched_contracts['剩余本金'].fillna(launched_contracts['放款金额'])

        # 2.3 剩余本金>0才计入在贷
        contract_principal = launched_contracts[launched_contracts['剩余本金'] > 0]['剩余本金']

        if contract_principal.empty:
            balance_series = pd.Series(0, index=labels, name=month_str)
            monthly_balance_list.append(balance_series)
            continue

        # 3. 计算截至该月末的时点DPD（仅统计：已到期 + 截至月末仍未结清 的期数）
        mask_due_unsettled = (
                (df_real['租金结算日期'] <= month_end)
                & ((df_real['结清日期'].isna()) | (df_real['结清日期'] > month_end))
                & df_real['合同号'].isin(contract_principal.index)
        )
        df_due_unsettled = df_real[mask_due_unsettled].copy()

        if not df_due_unsettled.empty:
            df_due_unsettled['时点DPD'] = (month_end - df_due_unsettled['租金结算日期']).dt.days
            df_due_unsettled['时点DPD'] = df_due_unsettled['时点DPD'].clip(lower=0)
            contract_dpd = df_due_unsettled.groupby('合同号')['时点DPD'].max()
        else:
            contract_dpd = pd.Series(dtype='float64')

        # 4. 合并DPD与本金，未到期/无逾期合同DPD填0（归为M0）
        contract_status = pd.concat([contract_dpd, contract_principal], axis=1)
        contract_status.columns = ['时点DPD', '未偿还本金']
        contract_status['时点DPD'] = contract_status['时点DPD'].fillna(0)

        # 5. 划分逾期等级
        contract_status['逾期等级'] = pd.cut(
            contract_status['时点DPD'],
            bins=bins,
            labels=labels,
            right=False,
            include_lowest=True
        )

        # 6. 汇总各等级余额
        balance = contract_status.groupby('逾期等级', observed=False)['未偿还本金'].sum()
        balance.name = month_str
        monthly_balance_list.append(balance)

        # ========== 新增：导出指定月份的明细 ==========
        if export_month and month_str == export_month:
            # 合并合同基础信息，方便核对
            export_detail = contract_status.join(contract_base[['放款金额', '起租日']])
            # 重置索引，把合同号变成列
            export_detail = export_detail.reset_index()
            # 按剩余本金降序排列，方便核对大额合同
            export_detail = export_detail.sort_values('未偿还本金', ascending=False)

            # 合并成最终余额表
    monthly_balance_df = pd.concat(monthly_balance_list, axis=1).fillna(0)
    monthly_balance_df.loc['合计'] = monthly_balance_df.sum()

    # 导出明细
    if export_detail is not None:
        export_detail.to_excel(export_path, index=False)
        logger.info(f"{export_month} 月末合同明细已导出至：{export_path}，共 {len(export_detail)} 条在贷合同")
        logger.info(f"当月计算总余额：{export_detail['未偿还本金'].sum():,.2f} 元")

    logger.info(f"月度余额表生成完成，共覆盖 {len(month_range)} 个观察月")
    return monthly_balance_df


# ============================================
# 7. 迁移率表（对应教程表2）
# ============================================
def build_roll_rate_table(monthly_balance_df, fixed_m7_recovery_rate=None):
    """
    基于月度余额表计算迁移率，完全对应教程「表2-迁移率统计」
    参数：
        monthly_balance_df: 月度余额表，行=逾期等级，列=月份
        m7_recovery_series: Series，索引为月份(YYYY-MM)，值为当月M7回收金额
    """
    pairs = CONFIG["business_params"]["roll_rate_pairs"]
    months = monthly_balance_df.columns.tolist()

    if len(months) < 2:
        logger.warning("观察月份不足2个，无法计算迁移率")
        return pd.DataFrame()

    roll_rate_data = {}
    valid_months = months[1:]  # 迁徙率有效月份：从第2个月开始（有上月数据才能计算迁徙率）

    # 1. 计算各档迁移率：本月Mn+1余额 / 上月Mn余额
    for prev_label, curr_label in pairs:
        rate_name = f"{prev_label.split('(')[0]}-{curr_label.split('(')[0]}"
        monthly_rates = {}
        for i in range(1, len(months)):
            prev_month = months[i - 1]
            curr_month = months[i]
            prev_bal = monthly_balance_df.loc[prev_label, prev_month]
            curr_bal = monthly_balance_df.loc[curr_label, curr_month]
            raw_rate = curr_bal / prev_bal if prev_bal != 0 else 0
            monthly_rates[curr_month] = pd.Series([raw_rate]).clip(0, 1).iloc[0]
        roll_rate_data[rate_name] = monthly_rates

    # 2. 若传入固定回收率，则全月统一填充该值作为M7回收率
    if fixed_m7_recovery_rate is not None:
        recovery_rates = {month: fixed_m7_recovery_rate for month in valid_months}
        roll_rate_data['M7假设回收率'] = recovery_rates

    # 3. 转DataFrame并计算平均值
    roll_rate_df = pd.DataFrame(roll_rate_data).T
    roll_rate_df['近12月平均值'] = roll_rate_df.apply(lambda x: x.dropna().tail(12).mean(), axis=1)

    # 4. 调整列顺序：平均值放最前
    cols = ['近12月平均值'] + [c for c in roll_rate_df.columns if c != '近12月平均值']
    roll_rate_df = roll_rate_df[cols]

    logger.info("迁移率表计算完成")
    return roll_rate_df


# ============================================
# 8. 各级坏账损失率计算（仅用近12月平均迁徙率）
# ============================================
def calc_loss_rates(roll_rate_df, recovery_rate=0.3):
    """
    基于近12月平均迁徙率，计算M0~M7每一档的累计毛/净坏账损失率
    异常处理：单档迁徙率超过100%时，按100%封顶
    """
    # 严格按逾期递进顺序定义迁徙链条
    roll_chain = ['M0-M1', 'M1-M2', 'M2-M3', 'M3-M4', 'M4-M5', 'M5-M6', 'M6-M7']
    level_labels = ['M0', 'M1', 'M2', 'M3', 'M4', 'M5', 'M6']

    # 校验迁徙率档位是否齐全
    missing = [x for x in roll_chain if x not in roll_rate_df.index]
    if missing:
        logger.error(f"迁徙率缺少档位，无法计算损失率：{missing}")
        return pd.DataFrame()

    # ========== 测试日志1：打印迁徙率索引，确认名称匹配 ==========
    logger.info(f"迁徙率表行索引：{roll_rate_df.index.tolist()}")
    logger.info(f"待匹配迁徙链条：{roll_chain}")

    # 仅提取近12月平均值，简化计算
    avg_roll = roll_rate_df.loc[roll_chain, '近12月平均值'].copy()

    # 异常值处理：迁徙率封顶100%，最低0%
    avg_roll = avg_roll.clip(lower=0, upper=1)
    logger.info(f"裁剪后平均迁徙率：\n{avg_roll}")

    # 核心：从后往前累乘，得到每一档到M7的累计毛损失率
    gross_loss_series = avg_roll[::-1].cumprod()[::-1]
    gross_loss_series.index = level_labels

    # ========== 新增：补全M7等级损失率 ==========
    # M7已为呆账，毛损失率100%
    gross_loss_series['M7'] = 1.0

    # 计算净损失率：毛损失率 × (1-回收率)
    net_loss_series = gross_loss_series * (1 - recovery_rate)

    # 整理为标准DataFrame，按逾期等级排序
    loss_rates = pd.DataFrame({
        '毛损失率': gross_loss_series,
        '净损失率': net_loss_series
    })
    loss_rates.index.name = '逾期等级'
    # 强制按M0→M7顺序排列
    loss_rates = loss_rates.reindex(['M0', 'M1', 'M2', 'M3', 'M4', 'M5', 'M6', 'M7'])

    logger.info("各级坏账损失率计算完成（含M7，仅近12月平均口径）")
    logger.info(f"损失率表内容：\n{loss_rates}")
    return loss_rates


# ============================================
# 9. 月度拨备准备金测算
# ============================================
def calc_monthly_provision(monthly_balance_df, loss_rates_df):
    """
    基于月度余额表和各级净损失率，计算每月应计提的坏账准备金
    自动对齐余额表的带括号全称标签
    """
    # 提取各级净损失率（简称索引）
    net_loss = loss_rates_df['净损失率'].copy()

    # ========== 新增：索引映射，对齐余额表的全称标签 ==========
    # 与配置中的 roll_rate_labels 严格对应，顺序 M0→M7
    label_mapping = {
        'M0': 'M0(正常)',
        'M1': 'M1(1-30天)',
        'M2': 'M2(31-60天)',
        'M3': 'M3(61-90天)',
        'M4': 'M4(91-120天)',
        'M5': 'M5(121-150天)',
        'M6': 'M6(151-180天)',
        'M7': 'M7(181天+)'
    }
    # 替换损失率的索引为余额表的全称
    net_loss.index = net_loss.index.map(label_mapping)
    logger.info(f"映射后净损失率索引：{net_loss.index.tolist()}")

    # 提取余额表的逾期等级部分（排除合计行）
    balance = monthly_balance_df.drop(index='合计', errors='ignore')
    logger.info(f"余额表行索引：{balance.index.tolist()}")

    # 校验逾期等级是否对齐
    common_levels = balance.index.intersection(net_loss.index)
    logger.info(f"匹配成功的逾期等级：{common_levels.tolist()}")

    if len(common_levels) == 0:
        logger.error("余额表与损失率的逾期等级仍无法匹配，无法计算拨备")
        return pd.DataFrame(), pd.DataFrame()

    balance = balance.loc[common_levels]
    net_loss = net_loss.loc[common_levels]

    # 计算各级每月期望损失：当月余额 × 对应等级的净损失率
    provision_detail = balance.multiply(net_loss, axis=0)
    logger.info(f"拨备明细矩阵形状：{provision_detail.shape}，列数：{len(provision_detail.columns)}")

    # 计算月度汇总：总拨备、总资产、拨备率
    monthly_provision = pd.DataFrame({
        '月末总资产余额': balance.sum(axis=0),
        '当月应计提拨备': provision_detail.sum(axis=0)
    })
    monthly_provision['应计提拨备率'] = monthly_provision['当月应计提拨备'] / monthly_provision['月末总资产余额']

    logger.info(f"月度拨备汇总形状：{monthly_provision.shape}")
    logger.info(f"月度拨备汇总前3行：\n{monthly_provision.head(3)}")

    logger.info("月度拨备准备金测算完成")
    return monthly_provision, provision_detail


# ============================================
# 10. 静态池（Vintage）逾期余额表
# ============================================
def build_vintage_balance(df, analysis_date):
    """
    按起租批次（Vintage）+账龄（MOB）统计各逾期等级余额，为静态池迁徙率做准备
    输出：多级索引 [Vintage, MOB]，列=各逾期等级，值=对应月末本金余额
    """
    bins = CONFIG["business_params"]["roll_rate_bins"]
    labels = CONFIG["business_params"]["roll_rate_labels"]

    df_real = df.copy()
    # 提取合同级基础信息（起租日、放款金额），Vintage已在预处理阶段生成
    contract_base = df_real.groupby('合同号').agg(
        放款金额=('放款金额', 'first'),
        起租日=('起租日', 'min'),
        Vintage=('Vintage', 'first')
    ).reset_index()
    contract_base = contract_base.set_index('合同号')
    # 同口径：仅保留2023年及以后起租
    contract_base = contract_base[contract_base['起租日'] >= pd.Timestamp('2023-01-01')]

    # 生成观察月份范围
    min_date = contract_base['起租日'].min()
    if pd.isna(min_date):
        logger.error("无有效起租日期，无法生成静态池余额表")
        return pd.DataFrame()

    month_range = pd.period_range(start=min_date.to_period('M'), end=analysis_date.to_period('M'), freq='M')
    month_ends = [pd.Timestamp(m.end_time) for m in month_range]

    # 存储所有批次所有账龄的余额数据
    vintage_balance_list = []

    for month_end in month_ends:
        month_str = month_end.strftime('%Y-%m')
        # 筛选截至当月已起租的合同
        mask_launched = contract_base['起租日'] <= month_end
        launched_contracts = contract_base[mask_launched].copy()

        if launched_contracts.empty:
            continue

        # 计算当月剩余本金（完全复用动态池逻辑，口径一致）
        mask_settled = (
                df_real['结清日期'].notna()
                & (df_real['结清日期'] <= month_end)
                & df_real['合同号'].isin(launched_contracts.index)
        )
        df_settled = df_real[mask_settled]

        if not df_settled.empty:
            df_settled_sorted = df_settled.sort_values('期号')
            settled_principal = df_settled_sorted.groupby('合同号')['未偿还本金'].last()
        else:
            settled_principal = pd.Series(dtype='float64')

        launched_contracts['剩余本金'] = settled_principal
        launched_contracts['剩余本金'] = launched_contracts['剩余本金'].fillna(launched_contracts['放款金额'])
        contract_principal = launched_contracts[launched_contracts['剩余本金'] > 0]['剩余本金']

        if contract_principal.empty:
            continue

        # 计算当月时点DPD与逾期等级（完全复用动态池逻辑）
        mask_due_unsettled = (
                (df_real['租金结算日期'] <= month_end)
                & ((df_real['结清日期'].isna()) | (df_real['结清日期'] > month_end))
                & df_real['合同号'].isin(contract_principal.index)
        )
        df_due_unsettled = df_real[mask_due_unsettled].copy()

        if not df_due_unsettled.empty:
            df_due_unsettled['时点DPD'] = (month_end - df_due_unsettled['租金结算日期']).dt.days
            df_due_unsettled['时点DPD'] = df_due_unsettled['时点DPD'].clip(lower=0)
            contract_dpd = df_due_unsettled.groupby('合同号')['时点DPD'].max()
        else:
            contract_dpd = pd.Series(dtype='float64')

        # 合并合同状态
        contract_status = pd.concat([contract_dpd, contract_principal], axis=1)
        contract_status.columns = ['时点DPD', '未偿还本金']
        contract_status['时点DPD'] = contract_status['时点DPD'].fillna(0)

        # 划分逾期等级
        contract_status['逾期等级'] = pd.cut(
            contract_status['时点DPD'],
            bins=bins,
            labels=labels,
            right=False,
            include_lowest=True
        )

        # ========== 静态池核心：关联Vintage + 计算MOB ==========
        # 关联每个合同的Vintage
        contract_status = contract_status.join(contract_base['Vintage'])
        # 计算当月对应的账龄MOB：月末月份 - 起租月份
        current_period = month_end.to_period('M')
        vintage_periods = pd.PeriodIndex(contract_status['Vintage'], freq='M')
        contract_status['MOB'] = (current_period - vintage_periods).map(lambda x: x.n)

        # 按 Vintage + MOB + 逾期等级 汇总余额
        monthly_vintage_bal = (
            contract_status.groupby(['Vintage', 'MOB', '逾期等级'], observed=False)['未偿还本金']
            .sum()
            .unstack(level='逾期等级')
            .fillna(0)
        )

        vintage_balance_list.append(monthly_vintage_bal)

    # 合并所有月份的结果，去重（同一个Vintage的同一个MOB只会出现一次）
    vintage_balance_df = pd.concat(vintage_balance_list)
    vintage_balance_df = vintage_balance_df[~vintage_balance_df.index.duplicated(keep='last')]
    vintage_balance_df = vintage_balance_df.sort_index()

    logger.info(
        f"静态池余额表生成完成，共 {vintage_balance_df.index.get_level_values('Vintage').nunique()} 个批次，最大账龄MOB={vintage_balance_df.index.get_level_values('MOB').max()}")
    return vintage_balance_df


# ============================================
# 11. 静态池（Vintage）迁徙率计算
# ============================================
def build_vintage_roll_rate(vintage_balance_df):
    """
    基于静态池余额表，计算各Vintage的分MOB迁徙率，以及跨批次平均基准曲线
    返回：
      - vintage_roll_detail：各Vintage分MOB的迁徙率明细（多级索引）
      - avg_roll_curve：跨批次平均的迁徙率曲线（行=迁徙档位，列=MOB）
    """
    pairs = CONFIG["business_params"]["roll_rate_pairs"]
    # 提取所有Vintage和MOB
    all_vintages = vintage_balance_df.index.get_level_values('Vintage').unique().tolist()
    all_mobs = sorted(vintage_balance_df.index.get_level_values('MOB').unique())

    detail_list = []

    # 遍历每个Vintage，计算内部相邻MOB的迁徙率
    for vintage in all_vintages:
        # 取出当前批次的各MOB余额
        batch_balance = vintage_balance_df.loc[vintage]
        batch_mobs = sorted(batch_balance.index.tolist())

        if len(batch_mobs) < 2:
            continue

        batch_rates = {}
        # 遍历相邻的MOB，计算各档迁徙率
        for i in range(1, len(batch_mobs)):
            prev_mob = batch_mobs[i - 1]
            curr_mob = batch_mobs[i]
            mob_label = f"MOB{prev_mob}→MOB{curr_mob}"

            rate_dict = {}
            for prev_label, curr_label in pairs:
                prev_bal = batch_balance.loc[prev_mob, prev_label] if prev_label in batch_balance.columns else 0
                curr_bal = batch_balance.loc[curr_mob, curr_label] if curr_label in batch_balance.columns else 0
                rate_name = f"{prev_label.split('(')[0]}-{curr_label.split('(')[0]}"
                rate_dict[rate_name] = curr_bal / prev_bal if prev_bal != 0 else pd.NA

            batch_rates[mob_label] = rate_dict

        # 转为DataFrame，加入Vintage标识
        batch_df = pd.DataFrame(batch_rates).T
        batch_df.index.name = '账龄跃迁'
        batch_df['Vintage'] = vintage
        detail_list.append(batch_df.reset_index().set_index(['Vintage', '账龄跃迁']))

    # 合并所有批次明细
    if not detail_list:
        logger.warning("无足够数据计算静态池迁徙率")
        return pd.DataFrame(), pd.DataFrame()

    vintage_roll_detail = pd.concat(detail_list)
    vintage_roll_detail = vintage_roll_detail.sort_index()

    # 计算跨批次平均基准曲线（按账龄跃迁取平均，排除空值）
    avg_roll_curve = vintage_roll_detail.groupby('账龄跃迁').mean(numeric_only=True).T
    # 异常值兜底：迁徙率封顶100%
    avg_roll_curve = avg_roll_curve.clip(lower=0, upper=1)

    # 补充每个MOB的样本批次数量，用于判断可信度
    sample_count = vintage_roll_detail.groupby('账龄跃迁').count().iloc[:, 0]
    avg_roll_curve.loc['样本批次数量'] = sample_count

    logger.info("静态池迁徙率计算完成")
    return vintage_roll_detail, avg_roll_curve


if __name__ == "__main__":
    try:
        # 1. 确定分析基准日（上月最后一天）
        today = pd.Timestamp.now()  # 修正：保留pd.Timestamp类型，不转为字符串
        last_day_of_last_month = today.replace(day=1) - pd.Timedelta(days=1)  # 计算上月最后一天（日期类型）
        ANALYSIS_DATE = last_day_of_last_month  # 此时ANALYSIS_DATE是pd.Timestamp类型
        logger.info(f"分析基准日自动设置为：{ANALYSIS_DATE.strftime('%Y-%m-%d')}")  # 日志打印时再转为字符串（不影响计算）

        # 1.1 数据加载与合并
        df_merged = merge_excel_files(ANALYSIS_DATE)
        if df_merged is None:
            raise ValueError("未成功加载数据")

        # 2. 数据预处理
        filter_params = {
            "经销商名称": CONFIG["dealer_filters"],
            "业务模式": ["损失不担模式"],
            "业务类别": ["商用车"]  # "大区": ["华北大区"]
        }
        df_clean_base = preprocess_data(df_merged, filter_params)
        df_clean = fill_missing_periods(df_clean_base)

        # 3. 构建Vintage矩阵
        vintage_mob_matrix = build_vintage_matrix(df_clean)

        # 5. 过滤无效MOB期数
        vintage_mob_valid = filter_invalid_mob(vintage_mob_matrix, ANALYSIS_DATE)

        # 计算迁徙率，生成月度余额表（表1）
        monthly_balance = build_monthly_balance(df_clean_base, ANALYSIS_DATE, export_month='2026-06',
                                                export_path='202606月末余额核对明细.xlsx')

        # 生成迁移率表（表2）
        fixed_m7_rate = 0.3
        roll_rate_table = build_roll_rate_table(monthly_balance, fixed_m7_recovery_rate=fixed_m7_rate)

        vintage_balance = build_vintage_balance(df_clean_base, ANALYSIS_DATE)  # 1. 生成静态池余额立方体
        vintage_roll_detail, avg_roll_curve = build_vintage_roll_rate(vintage_balance)  # 2. 计算静态池迁徙率（明细+平均曲线）
        loss_rates = calc_loss_rates(roll_rate_table, recovery_rate=fixed_m7_rate)  # 1. 计算M0~M6各级毛/净损失率
        monthly_provision, provision_detail = calc_monthly_provision(monthly_balance,
                                                                     loss_rates)  # 2. 计算月度拨备（汇总表 + 各级明细）

        # 6. 生成透视表
        pivot_table = vintage_mob_valid.pivot_table(
            index='Vintage',
            columns='期号',
            values='逾期率',
            aggfunc='mean'
        ).fillna('-')

        # 7. 保存结果到Excel
        output_path = CONFIG["output_file"]
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # 保存Vintage矩阵透视表
            pivot_table.to_excel(writer, sheet_name='Vintage矩阵', index_label='Vintage')

            # 保存有效明细数据
            vintage_mob_valid.to_excel(writer, sheet_name='有效明细数据', index=False)

            # 保存首次逾期明细（仅保留每个合同第一次逾期的记录）
            # 1. 先筛选出所有逾期记录（is_overdue=1）
            overdue_records = df_clean[
                df_clean['is_overdue'] == 1  # 只保留逾期记录
                ][
                [
                    'Vintage', '合同号', '经销商名称', '客户名称', '大区', '业务来源', '起租日', '期号',
                    '放款金额', '未偿还本金', '逾期天数（DPD）', '首次逾期期号', '首次逾期未偿还本金'
                ]
            ]

            # 2. 按合同号分组，每组只保留期号最小的那条记录（即首次逾期）
            first_overdue_only = overdue_records.sort_values(by=['合同号', '期号']) \
                .groupby('合同号').first().reset_index()

            # 3. 按Vintage和期号排序，确保展示有序
            first_overdue_only = first_overdue_only[
                first_overdue_only['首次逾期未偿还本金'] > 0.01]  # 过滤结清场景与最后一期逾期场景
            first_overdue_only = first_overdue_only.sort_values(by=['Vintage', '期号', '合同号'])

            # 写入Excel（工作表名称更改为“首次逾期明细”）
            first_overdue_only.to_excel(writer, sheet_name='首次逾期明细', index=False)

            # ========== 新增：写入迁徙率相关Sheet ==========
            if not monthly_balance.empty:
                monthly_balance.to_excel(writer, sheet_name='表1-月度余额分布')

            if not roll_rate_table.empty:
                roll_rate_table.to_excel(writer, sheet_name='表2-迁移率统计', index_label='月度迁移率')
                # 设置百分比格式
                ws_roll = writer.sheets['表2-迁移率统计']
                for row in ws_roll.iter_rows(min_row=2, min_col=2):
                    for cell in row:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '0.00%'

            # if not df_clean_base.empty:
            # df_clean_base.to_excel(writer, sheet_name='表3-租金明细表')

            # 写入损失率表
            if not loss_rates.empty:
                loss_rates.to_excel(writer, sheet_name='表3-各级损失率')

            # 写入拨备明细（各级别每月期望损失）
            if not provision_detail.empty:
                provision_detail.to_excel(writer, sheet_name='表4-拨备明细矩阵')

            # 写入月度拨备汇总
            if not monthly_provision.empty:
                monthly_provision.to_excel(writer, sheet_name='表5-月度拨备汇总')

            # 新增静态池相关Sheet
            if not vintage_balance.empty:
                vintage_balance.to_excel(writer, sheet_name='静态池余额明细')
            if not vintage_roll_detail.empty:
                vintage_roll_detail.to_excel(writer, sheet_name='静态池迁徙率明细')
            if not avg_roll_curve.empty:
                avg_roll_curve.to_excel(writer, sheet_name='静态池平均迁徙曲线')

            # 美化Vintage矩阵格式
            worksheet = writer.sheets['Vintage矩阵']
            valid_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            for row in worksheet.iter_rows(min_row=2, min_col=2):
                for cell in row:
                    if cell.value != '-':
                        cell.fill = valid_fill

        logger.info(f"分析完成，结果已保存至: {output_path}")
        print(f"分析完成，结果已保存至: {output_path}")

    except Exception as e:
        logger.error(f"主程序执行错误：{e}", exc_info=True)
