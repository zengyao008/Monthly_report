# 合并文件夹内的表格，并去重（用于客户信息一览表）
import os
import pandas as pd
from openpyxl import load_workbook

# ====================== 【1. 你的文件夹路径（不用改）】 ======================
folder_path = r"D:\Filelist\Python_list\File_list\merge_list01"

# ====================== 【2. 目标文件+工作表配置（按你的需求填好）】 ======================
target_excel_path = r"C:\Users\zengy\Desktop\通盛个人文件\7.评审部文件\1.数据存档\时效台账-总表.xlsx"
target_sheet_name = "商用车客户信息一览表"

# 存储所有数据
all_data = []

# 遍历文件夹里的所有文件
for filename in os.listdir(folder_path):
    # 只处理 Excel 文件（.xlsx 和 .xls）
    if filename.endswith((".xlsx", ".xls")):
        file_path = os.path.join(folder_path, filename)
        print(f"正在处理：{filename}")

        try:
            # 读取 Excel 所有数据
            df = pd.read_excel(file_path)
            all_data.append(df)

        except Exception as e:
            print(f"❌ 处理 {filename} 失败：{e}")

# ====================== 合并所有数据 + 清理（完全保留你原来的逻辑） ======================
if all_data:
    merged_df = pd.concat(all_data, ignore_index=True)
    merged_df = merged_df.drop_duplicates()  # 去重

    # ====================== 删除【全空列】 ======================
    print("\n🧹 开始清理全空列...")
    empty_cols = [col for col in merged_df.columns if merged_df[col].isnull().all()]
    total_cols_before = len(merged_df.columns)

    if empty_cols:
        merged_df = merged_df.drop(columns=empty_cols)
        print(f"✅ 已删除【{len(empty_cols)}】个全空列：")
        for col in empty_cols:
            print(f"   - {col}")
    else:
        print("✅ 未发现全空列")

    total_cols_after = len(merged_df.columns)
    print(f"\n📊 列数变化：{total_cols_before} 列 → {total_cols_after} 列")

    # ====================== 【核心修改：替换指定Excel的指定工作表】 ======================
    print(f"\n📝 正在写入目标Excel：{target_excel_path}")
    print(f"📋 目标工作表：{target_sheet_name}")

    try:
        # 1. 先判断目标Excel文件是否存在
        if os.path.exists(target_excel_path):
            # 加载已有工作簿
            book = load_workbook(target_excel_path)
            # 如果目标sheet已存在，先删除（避免旧数据残留）
            if target_sheet_name in book.sheetnames:
                del book[target_sheet_name]
                print(f"✅ 已清空旧的工作表：{target_sheet_name}")
            # 保存修改后的工作簿
            book.save(target_excel_path)

        # 2. 写入新数据到目标sheet
        with pd.ExcelWriter(
            target_excel_path,
            engine='openpyxl',
            mode='a',  # 追加模式，不影响其他sheet
            if_sheet_exists='replace'
        ) as writer:
            merged_df.to_excel(writer, sheet_name=target_sheet_name, index=False)

        # 3. 写入完成，输出结果
        print("\n✅ 数据写入完成！")
        print(f"📊 共合并 {len(all_data)} 个文件")
        print(f"📁 目标文件：{target_excel_path}")
        print(f"📋 目标工作表：{target_sheet_name}")
        print(f"🧾 写入数据总行数：{len(merged_df)}")
        print(f"📋 写入数据总列数：{total_cols_after}")

    except Exception as e:
        print(f"\n❌ 写入目标Excel失败：{e}")

else:
    print("\n❌ 没有找到任何可合并的数据")