# 程序可直接应用,还原资产余额表重组数据
import pandas as pd
import re
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# 抑制openpyxl的样式警告
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# 定义文件路径和文件名
folder_path = r"D:\Filelist\Python_list\历史系统数据"
file_name = "资产余额表06.30"


def find_file_with_extension(folder, name):
    """查找指定文件夹中名称匹配的文件（忽略扩展名）"""
    for ext in ['.xlsx']:  # 仅支持xlsx（openpyxl不支持xls）
        full_path = os.path.join(folder, f"{name}{ext}")
        if os.path.exists(full_path):
            return full_path
    # 若未找到xlsx，检查xls（需单独处理，但优先保证格式保留）
    for ext in ['.xls']:
        full_path = os.path.join(folder, f"{name}{ext}")
        if os.path.exists(full_path):
            raise ValueError(f"文件格式为.xls，建议先另存为.xlsx以保留格式（openpyxl不支持xls）")
    raise FileNotFoundError(f"在 {folder} 中未找到 {name}.xlsx")


def load_data(file_path):
    """读取Excel数据（仅用于处理逻辑）"""
    return pd.read_excel(file_path)


def process_data(dataset):
    """处理资产余额表数据（保持原逻辑，确保列数与原表一致）"""
    original_columns = dataset.columns.tolist()  # 严格保留原始列顺序

    # 标记重组合同
    dataset['是否重组合同'] = dataset['合同号'].str.contains(r'-(?:CZ|RA)', na=False)
    # 提取原合同号
    dataset['原合同号'] = dataset['合同号'].apply(
        lambda x: re.sub(r'-(?:CZ|RA).*$', '', x) if pd.notnull(x) else x
    )

    # 取原始合同（非重组+去重）
    original_all_fields = dataset[~dataset['是否重组合同']].drop_duplicates(
        subset=['原合同号'], keep='first'
    )

    # 分组计算最大值字段
    max_fields = ['合同到期日期', '剩余本金', '剩余期数', '逾期天数']
    for field in max_fields + ['合同号']:
        if field not in dataset.columns:
            raise ValueError(f"数据中缺少必要的列: {field}")
    grouped_max = dataset.groupby('原合同号')[max_fields].max().reset_index()

    # 合并原始合同与最大值（避免列名冲突）
    result = pd.merge(
        original_all_fields, grouped_max, on='原合同号', how='left', suffixes=('', '_max')
    )

    # 用最大值更新目标字段
    for field in max_fields:
        if f'{field}_max' in result.columns:
            result[field] = result[f'{field}_max']
            result = result.drop(columns=[f'{field}_max'])

    # 处理无原始合同的情况
    missing_originals = set(dataset['原合同号']) - set(original_all_fields['原合同号'])
    if missing_originals:
        补充数据 = dataset[dataset['原合同号'].isin(missing_originals)].drop_duplicates(
            subset=['原合同号'], keep='first'
        )
        补充数据 = pd.merge(补充数据, grouped_max, on='原合同号', how='left', suffixes=('', '_max'))
        for field in max_fields:
            if f'{field}_max' in 补充数据.columns:
                补充数据[field] = 补充数据[f'{field}_max']
                补充数据 = 补充数据.drop(columns=[f'{field}_max'])
        result = pd.concat([result, 补充数据], ignore_index=True)

    # 清理临时列+恢复原始列顺序（核心：确保列数、列名与原表完全一致）
    result = result.drop(columns=['是否重组合同', '原合同号'], errors='ignore')
    # 补全原表中存在但处理后可能缺失的列
    for col in original_columns:
        if col not in result.columns:
            result[col] = None  # 缺失列填充空值，保证列顺序一致
    result = result[original_columns]  # 强制按原表列顺序排列

    return result


def update_excel_with_format(original_file, processed_df, sheet_name="Sheet1"):
    """
    在保留原格式的前提下，用处理后的数据更新Excel文件
    """
    # 1. 加载原始Excel（含格式）
    wb = load_workbook(original_file)
    # 2. 获取目标工作表（默认Sheet1，可根据实际修改）
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"工作表 {sheet_name} 不存在，请检查原表")
    ws = wb[sheet_name]

    # 3. 清空原数据区域（保留表头和格式）
    # 假设第1行为表头，从第2行开始清空
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row > 1:  # 只清空数据行，保留表头
        for row in range(2, max_row + 1):
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).value = None  # 清空单元格值（保留格式）

    # 4. 写入处理后的数据（从第2行开始，覆盖旧数据）
    for r_idx, row in enumerate(dataframe_to_rows(processed_df, index=False, header=False), 2):
        for c_idx, value in enumerate(row, 1):
            # 只写入有效列（避免超出原表列数导致格式错乱）
            if c_idx <= max_col:
                ws.cell(row=r_idx, column=c_idx).value = value

    # 5. 另存为新文件（避免覆盖原始文件）
    output_file = os.path.splitext(original_file)[0] + "_处理后（保留格式）.xlsx"
    wb.save(output_file)
    return output_file


def main():
    try:
        # 查找原始文件（仅支持xlsx格式，确保格式可保留）
        file_path = find_file_with_extension(folder_path, file_name)
        print(f"找到文件: {file_path}")

        # 读取数据（仅用于处理逻辑）
        dataset = load_data(file_path)
        print(f"成功读取数据，共 {len(dataset)} 行，{len(dataset.columns)} 列")

        # 处理数据（确保列数、列名与原表一致）
        processed_data = process_data(dataset)
        print(f"数据处理完成，处理后共 {len(processed_data)} 行，{len(processed_data.columns)} 列")

        # 关键步骤：更新Excel并保留格式
        output_file = update_excel_with_format(
            original_file=file_path,
            processed_df=processed_data,
            sheet_name="资产余额表"  # 替换为你原表的工作表名（如"资产余额表"）
        )
        print(f"处理后的数据已保存至: {output_file}")
        print("提示：文件保留了原表格式，可直接用于Power BI分析")

    except Exception as e:
        print(f"处理时出错: {str(e)}")


if __name__ == "__main__":
    main()
